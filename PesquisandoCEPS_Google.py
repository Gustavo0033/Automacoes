from openpyxl import load_workbook
import subprocess
from selenium import webdriver as WB
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
import pyautogui as pgui
import xlsxwriter
import subprocess
from openpyxl import load_workbook

nomeArquivoExcel = '/Users/gustavomendonca/Documents/CEPs/CEPs2.xlsx'
carregarPlanilha = load_workbook(nomeArquivoExcel)

planilhaDados = carregarPlanilha['CEPs']

navegador = WB.Chrome()
navegador.get('https://buscacepinter.correios.com.br/app/endereco/index.php')

navegador.find_element(By.ID, 'endereco').send_keys('07024-060')
navegador.find_element(By.ID, 'endereco').send_keys(Keys.RETURN)
pgui.sleep(2)

for linha in range(2, len(planilhaDados['A'])+ 1):
    navegador.find_element(By.ID, 'btn_nbusca').click()
    pgui.sleep(3)

    cepPesquisa = planilhaDados['A%s' % linha].value
    navegador.find_element(By.NAME, 'endereco').send_keys(cepPesquisa)
    navegador.find_element(By.NAME, 'btn_pesquisar').click()
    pgui.sleep(3)

    rua = navegador.find_element(By.XPATH, '//*[@id="resultado-DNEC"]/tbody/tr/td[1]').text
    print(rua)
    bairro = navegador.find_element(By.XPATH, '//*[@id="resultado-DNEC"]/tbody/tr/td[2]').text
    print(bairro)
    cidade = navegador.find_element(By.XPATH, '//*[@id="resultado-DNEC"]/tbody/tr/td[3]').text
    print(cidade)
    cep = navegador.find_element(By.XPATH, '//*[@id="resultado-DNEC"]/tbody/tr/td[4]').text
    print(cep)
    pgui.sleep(3)


    planilhaEnderecos = carregarPlanilha['Dados']
    linhaCEP = len(planilhaEnderecos['A']) + 1

    colunaA = 'A' + str(linhaCEP)
    colunaB = 'B' + str(linhaCEP)
    colunaC = 'C' + str(linhaCEP)
    colunaD = 'D' + str(linhaCEP)

    planilhaEnderecos[colunaA] = rua
    planilhaEnderecos[colunaB] = bairro
    planilhaEnderecos[colunaC] = cidade
    planilhaEnderecos[colunaD] = cep


carregarPlanilha.save(filename=nomeArquivoExcel)
subprocess.run(['open', nomeArquivoExcel])



